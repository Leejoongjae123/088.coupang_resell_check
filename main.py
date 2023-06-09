import json
import re

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBar, FormatObject, Rule
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,colors, Color,PatternFill


import requests
from bs4 import BeautifulSoup
import pprint
import time
import random
import openpyxl
import datetime




def get_col_width_row_height(img_width, img_height):
    col_width = img_width*63.2/504.19
    row_height = img_height*225.35/298.96
    return (col_width, row_height)
def load_excel(fname):
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(1, no_row + 1):
        name = ws.cell(row=i, column=1).value
        if name == "" or name == None:
            print('데이타 더 이상 없음')
            break
        data_list.append(name)
    print(data_list)
    return data_list
def calculate_duration(start_date, end_date):
    format = "%m/%d"  # 날짜 형식 지정
    start_datetime = datetime.datetime.strptime(start_date, format)  # 시작 날짜를 datetime 객체로 변환
    end_datetime = datetime.datetime.strptime(end_date, format)  # 종료 날짜를 datetime 객체로 변환

    duration = end_datetime - start_datetime  # 날짜 간의 차이 계산
    # print(duration.days)
    return duration.days  # 일 수 반환
def getUrlList(keyword):
    dataList=[]
    cookies = {
        'PCID': '7502917043777569504480',
        'x-coupang-accept-language': 'ko-KR',
        'x-coupang-target-market': 'KR',
        '_fbp': 'fb.1.1684062714833.1964233407',
        'sid': '0cd4049032304be58573fada46c29598a51802bc',
        'bm_sz': '40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456',
        'clickCoach': 'yes',
        '_ga': 'GA1.2.1852651095.1684922212',
        '_gid': 'GA1.2.708582323.1684922212',
        'searchKeyword': '%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80',
        'searchKeywordType': '%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D',
        'overrideAbTestGroup': '%5B%5D',
        'MARKETID': '7502917043777569504480',
        'bm_mi': 'ADB92FE5913B40FC5409CA9E81DA620B~YAAQpeQ1FwhMFBSIAQAAq0R2TRP61gInl5atkg8ZN6y2osbd38CFAPZUGFdWAU4wptsZyqVLLjWdyYCQpNjR6l4gin91TvULaWx32Ot443V+tNbzU/pcawf04lihCy6KAthfG9EHKlDA/r3ennltv5Ze7C13p+isaGyKo0M5kx4UDvURA4Bxsh70k1caHqwExFszEex8cgZrQ1pPHysk4R7bVjTznAuX/p2vUZ9e2vkZr6xntFkf0YmzvcCE5/KdjZqXh3lMpCde6FPhXrd/jyoYL2iwVf6vLhrBvc2/Crsjq7vQnzrDSvkLsgY=~1',
        'cto_bundle': 'ZBO22V9SaG5yR3k1VDRodiUyQlEwSXdiMnI2NmhyTDZBUEEwV3JOJTJCVllrOEdGMmdLMVdOeUJPb1V5Z1MyS0ZSU05lRTNBZTJFYkJRU2slMkJRSVA4NTA3WkZ3b1NHd2syZzBBNFVweFFzOEIzZWt6djhabkZKNU9oUFpqJTJGTUM5dFhVaTlPaGNoYkZ0SVIySFZFQVgyZm1vN1RXJTJGR0xBJTNEJTNE',
        # 'ak_bmsc': 'A6245E458AADC07EDF98845AAAFA3C7C~000000000000000000000000000000~YAAQpeQ1Fx1NFBSIAQAAX0t2TRNeeWcsOrFN4OdF6K+Vpu7CK4u2FUsb/iGtnMlu1slKbWxbxNrn8vKwKNwcv8ra7IOU8EYMXiBgjnUBJMbgwnmiixxXtdtUT+qZ5jvZhKj/P7PiskS7oyiZ7y3jBb877/8j2ZRY2+bjtTQa6NbGR6PhPnSDoZMvQBIdhnqE/e9Lq/QFIPLM6XqHWhdEetgWG1lbSXovTb9D35KDCeGkH8BoXM93RsbZWcaqWSpHlamsDwCLkBDiuWBEvIEIPhV+WcPeN+E/7QZdjfgXAysq5GFomm1t94v0JxMnkWJU1ExY5XZvkh0yL1hQ3F3ezDu+RbMQhhiZYxWQ/Ybtdzz5eHL+IjIYAWbfdy02wxJ9MLAXfL8eL3gh3PiQuQaVFc0H1JzILO65NF1zXOQ=',
        'bm_sv': 'ABBAE37C723B86B41ACFEBCDB0962346~YAAQpeQ1FyNPFBSIAQAA6VZ2TRNuARYwaMyMmtWR6k0wtBn+js8H2Ic156L7yh+OQG0EfUfNbK3KZWjVuzHNbmogm5deg/sSKAZVrROE2gPA5MuNLi5icooCPOCoVmnMoRWz66LC7SoVbyjpcgANyvM7V+gHkLicPwtkGddHp/7AOCJf54myBlL6xzJubE64u87UJ0J2IHJvgabQysKR2ZPcJsfl9WZsSKCoNBnN6pjp0FCH5wgkDXVSfSqBR/TePA==~1',
        '_abck': 'D8A3879F550D48A325858C749ED5EFC6~0~YAAQX3XTF9NHfD6IAQAAhep3TQmZTkA+hlTKR7DWzYzeVThp5wXDUJkti189ffkiW1L6S5k+poawjkqPKkB/fKRKFCHgI0IMa/Tjrbx3dCDSKdsg0wY9jglohkLO7AKd+sjmnMDUiU8P7SPY/g/HEVMTPDWsjvMmP//ZsQ3VJXNn+1iPFDAPjFy1P0sfR8XRza+CItbtnSyDeFhF498vUyitqRGqUiefLxvVzzlv0+5PSCqH/PCELnSoPuT0ZTL6hdPgXPZnAKDlsvGSWG1toYMiXYjjDdwQpCOp7p57dQwrrFlOVHFU8F1kuWUI3gbtwAohAsGhaTJPYUE/2uSbhJUqcZ53qOs0uUhttgDjqMC6JTtxbPGI0MAJ5nGipmXpOPCBTIntKHsxz3kCd0ZsynwvCiNLADTaFNnEp1ajRa4Q+hCg+6Xp~-1~-1~-1',
        'baby-isWide': 'small',
    }

    headers = {
        'authority': 'www.coupang.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        # 'cookie': 'PCID=7502917043777569504480; x-coupang-accept-language=ko-KR; x-coupang-target-market=KR; _fbp=fb.1.1684062714833.1964233407; sid=0cd4049032304be58573fada46c29598a51802bc; bm_sz=40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456; clickCoach=yes; _ga=GA1.2.1852651095.1684922212; _gid=GA1.2.708582323.1684922212; searchKeyword=%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80; searchKeywordType=%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D; overrideAbTestGroup=%5B%5D; MARKETID=7502917043777569504480; bm_mi=ADB92FE5913B40FC5409CA9E81DA620B~YAAQpeQ1FwhMFBSIAQAAq0R2TRP61gInl5atkg8ZN6y2osbd38CFAPZUGFdWAU4wptsZyqVLLjWdyYCQpNjR6l4gin91TvULaWx32Ot443V+tNbzU/pcawf04lihCy6KAthfG9EHKlDA/r3ennltv5Ze7C13p+isaGyKo0M5kx4UDvURA4Bxsh70k1caHqwExFszEex8cgZrQ1pPHysk4R7bVjTznAuX/p2vUZ9e2vkZr6xntFkf0YmzvcCE5/KdjZqXh3lMpCde6FPhXrd/jyoYL2iwVf6vLhrBvc2/Crsjq7vQnzrDSvkLsgY=~1; cto_bundle=ZBO22V9SaG5yR3k1VDRodiUyQlEwSXdiMnI2NmhyTDZBUEEwV3JOJTJCVllrOEdGMmdLMVdOeUJPb1V5Z1MyS0ZSU05lRTNBZTJFYkJRU2slMkJRSVA4NTA3WkZ3b1NHd2syZzBBNFVweFFzOEIzZWt6djhabkZKNU9oUFpqJTJGTUM5dFhVaTlPaGNoYkZ0SVIySFZFQVgyZm1vN1RXJTJGR0xBJTNEJTNE; ak_bmsc=A6245E458AADC07EDF98845AAAFA3C7C~000000000000000000000000000000~YAAQpeQ1Fx1NFBSIAQAAX0t2TRNeeWcsOrFN4OdF6K+Vpu7CK4u2FUsb/iGtnMlu1slKbWxbxNrn8vKwKNwcv8ra7IOU8EYMXiBgjnUBJMbgwnmiixxXtdtUT+qZ5jvZhKj/P7PiskS7oyiZ7y3jBb877/8j2ZRY2+bjtTQa6NbGR6PhPnSDoZMvQBIdhnqE/e9Lq/QFIPLM6XqHWhdEetgWG1lbSXovTb9D35KDCeGkH8BoXM93RsbZWcaqWSpHlamsDwCLkBDiuWBEvIEIPhV+WcPeN+E/7QZdjfgXAysq5GFomm1t94v0JxMnkWJU1ExY5XZvkh0yL1hQ3F3ezDu+RbMQhhiZYxWQ/Ybtdzz5eHL+IjIYAWbfdy02wxJ9MLAXfL8eL3gh3PiQuQaVFc0H1JzILO65NF1zXOQ=; bm_sv=ABBAE37C723B86B41ACFEBCDB0962346~YAAQpeQ1FyNPFBSIAQAA6VZ2TRNuARYwaMyMmtWR6k0wtBn+js8H2Ic156L7yh+OQG0EfUfNbK3KZWjVuzHNbmogm5deg/sSKAZVrROE2gPA5MuNLi5icooCPOCoVmnMoRWz66LC7SoVbyjpcgANyvM7V+gHkLicPwtkGddHp/7AOCJf54myBlL6xzJubE64u87UJ0J2IHJvgabQysKR2ZPcJsfl9WZsSKCoNBnN6pjp0FCH5wgkDXVSfSqBR/TePA==~1; _abck=D8A3879F550D48A325858C749ED5EFC6~0~YAAQX3XTF9NHfD6IAQAAhep3TQmZTkA+hlTKR7DWzYzeVThp5wXDUJkti189ffkiW1L6S5k+poawjkqPKkB/fKRKFCHgI0IMa/Tjrbx3dCDSKdsg0wY9jglohkLO7AKd+sjmnMDUiU8P7SPY/g/HEVMTPDWsjvMmP//ZsQ3VJXNn+1iPFDAPjFy1P0sfR8XRza+CItbtnSyDeFhF498vUyitqRGqUiefLxvVzzlv0+5PSCqH/PCELnSoPuT0ZTL6hdPgXPZnAKDlsvGSWG1toYMiXYjjDdwQpCOp7p57dQwrrFlOVHFU8F1kuWUI3gbtwAohAsGhaTJPYUE/2uSbhJUqcZ53qOs0uUhttgDjqMC6JTtxbPGI0MAJ5nGipmXpOPCBTIntKHsxz3kCd0ZsynwvCiNLADTaFNnEp1ajRa4Q+hCg+6Xp~-1~-1~-1; baby-isWide=small',
        'referer': 'https://www.coupang.com/',
        'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
    }

    params = {
        'component': '',
        'q': str(keyword),
        'channel': 'user',
    }

    response = requests.get('https://www.coupang.com/np/search', params=params, cookies=cookies, headers=headers)
    # print(response.text)
    soup=BeautifulSoup(response.text,'lxml')
    # print(soup.prettify())
    isUlTag=len(soup.find_all('ul',attrs={'id':'productList'}))
    if isUlTag==0:
        print("상품없음")
        dataList = []
    else:
        ulTag = soup.find('ul', attrs={'id': 'productList'})
        liTags = ulTag.find_all('li', attrs={'class': 'search-product'})
        for index,liTag in enumerate(liTags):

            isNumber=len(liTag.find_all('span',attrs={'class':re.compile("number no-\d+")}))
            if isNumber>=1:
                # print("찾았다!")
                name=liTag.find('div',attrs={'class':'name'}).get_text()
                url="https://www.coupang.com"+liTag.find('a',attrs={'class':'search-product-link'})['href']
                # print('url:',url)
                regex=re.compile("itemId=\d+")
                itemId=regex.findall(url)[0].replace("itemId=","")
                regex=re.compile("\d+")
                productNo=regex.findall(url)[0]
                regex = re.compile("vendorItemId=\d+")
                vendorItemId=regex.findall(url)[0].replace("vendorItemId=","")

                # regex=re.compile("/\d+")
                # productNo=regex.findall(url)[0].replace("/","")
                data={'name':name,'itemId':itemId,'vendorItemId':vendorItemId,'url':url,'productNo':productNo}
                # print(data)

                dataList.append(data)
        print(dataList)
    return dataList
def getCompanyInfo(data,prevVendorName):
    cookies = {
        'PCID': '7502917043777569504480',
        'x-coupang-accept-language': 'ko-KR',
        'x-coupang-target-market': 'KR',
        '_fbp': 'fb.1.1684062714833.1964233407',
        'sid': '0cd4049032304be58573fada46c29598a51802bc',
        'bm_sz': '40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456',
        'clickCoach': 'yes',
        '_ga': 'GA1.2.1852651095.1684922212',
        '_gid': 'GA1.2.708582323.1684922212',
        'searchKeyword': '%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80',
        'searchKeywordType': '%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D',
        'overrideAbTestGroup': '%5B%5D',
        'MARKETID': '7502917043777569504480',
        'bm_mi': 'ADB92FE5913B40FC5409CA9E81DA620B~YAAQpeQ1FwhMFBSIAQAAq0R2TRP61gInl5atkg8ZN6y2osbd38CFAPZUGFdWAU4wptsZyqVLLjWdyYCQpNjR6l4gin91TvULaWx32Ot443V+tNbzU/pcawf04lihCy6KAthfG9EHKlDA/r3ennltv5Ze7C13p+isaGyKo0M5kx4UDvURA4Bxsh70k1caHqwExFszEex8cgZrQ1pPHysk4R7bVjTznAuX/p2vUZ9e2vkZr6xntFkf0YmzvcCE5/KdjZqXh3lMpCde6FPhXrd/jyoYL2iwVf6vLhrBvc2/Crsjq7vQnzrDSvkLsgY=~1',
        # 'ak_bmsc': 'A6245E458AADC07EDF98845AAAFA3C7C~000000000000000000000000000000~YAAQpeQ1Fx1NFBSIAQAAX0t2TRNeeWcsOrFN4OdF6K+Vpu7CK4u2FUsb/iGtnMlu1slKbWxbxNrn8vKwKNwcv8ra7IOU8EYMXiBgjnUBJMbgwnmiixxXtdtUT+qZ5jvZhKj/P7PiskS7oyiZ7y3jBb877/8j2ZRY2+bjtTQa6NbGR6PhPnSDoZMvQBIdhnqE/e9Lq/QFIPLM6XqHWhdEetgWG1lbSXovTb9D35KDCeGkH8BoXM93RsbZWcaqWSpHlamsDwCLkBDiuWBEvIEIPhV+WcPeN+E/7QZdjfgXAysq5GFomm1t94v0JxMnkWJU1ExY5XZvkh0yL1hQ3F3ezDu+RbMQhhiZYxWQ/Ybtdzz5eHL+IjIYAWbfdy02wxJ9MLAXfL8eL3gh3PiQuQaVFc0H1JzILO65NF1zXOQ=',
        '_abck': 'D8A3879F550D48A325858C749ED5EFC6~0~YAAQX3XTF9NHfD6IAQAAhep3TQmZTkA+hlTKR7DWzYzeVThp5wXDUJkti189ffkiW1L6S5k+poawjkqPKkB/fKRKFCHgI0IMa/Tjrbx3dCDSKdsg0wY9jglohkLO7AKd+sjmnMDUiU8P7SPY/g/HEVMTPDWsjvMmP//ZsQ3VJXNn+1iPFDAPjFy1P0sfR8XRza+CItbtnSyDeFhF498vUyitqRGqUiefLxvVzzlv0+5PSCqH/PCELnSoPuT0ZTL6hdPgXPZnAKDlsvGSWG1toYMiXYjjDdwQpCOp7p57dQwrrFlOVHFU8F1kuWUI3gbtwAohAsGhaTJPYUE/2uSbhJUqcZ53qOs0uUhttgDjqMC6JTtxbPGI0MAJ5nGipmXpOPCBTIntKHsxz3kCd0ZsynwvCiNLADTaFNnEp1ajRa4Q+hCg+6Xp~-1~-1~-1',
        'FUN': '"{\'search\':[{\'reqUrl\':\'/search.pang\',\'isValid\':true}]}"',
        'cto_bundle': 'nGPxEl9SaG5yR3k1VDRodiUyQlEwSXdiMnI2Nm1rbFFRdEFrdlRWUEd5UEdndVRoQ0lVOCUyRkxaaExiTWFyWDlxcnFWS3RUNiUyRnFqVlRZMFRyNyUyRkRvczczS1hBdEdFZUxEdlJsazklMkJYekZDTG1GOXpha3YzZSUyQjNUWiUyQk9WU2FreGRSd01QVk0ySHJyb1pTTWFmSGV1aiUyRjBNNDlwRFpnJTNEJTNE',
        'baby-isWide': 'small',
        'bm_sv': 'ABBAE37C723B86B41ACFEBCDB0962346~YAAQLQ3VF5d/DxWIAQAAZeF6TRP+Ox+yJNm/qqLY7GCh17/bXQvj5sAPtReZTV2iDvIP5QvFy7tJROBe6fuk6uzFRhKKaRp+RvOCJ/dvxf01e6GMYmjX0erdpGLs2JdJWL0RNzuelix7uTdtMhtPvOuORz1e6t7/WMPKIdggQ8Lp3jBxSBeHXms46UXzU47pn77wzBBxqUdC9U5tSOe2oK4rNDPzMCc2Z7bRRbXPWL/PYTgXgQK1sCnae37KkIT45Fo=~1',
    }

    headers = {
        'authority': 'www.coupang.com',
        'accept': '*/*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        # 'cookie': 'PCID=7502917043777569504480; x-coupang-accept-language=ko-KR; x-coupang-target-market=KR; _fbp=fb.1.1684062714833.1964233407; sid=0cd4049032304be58573fada46c29598a51802bc; bm_sz=40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456; clickCoach=yes; _ga=GA1.2.1852651095.1684922212; _gid=GA1.2.708582323.1684922212; searchKeyword=%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80; searchKeywordType=%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D; overrideAbTestGroup=%5B%5D; MARKETID=7502917043777569504480; bm_mi=ADB92FE5913B40FC5409CA9E81DA620B~YAAQpeQ1FwhMFBSIAQAAq0R2TRP61gInl5atkg8ZN6y2osbd38CFAPZUGFdWAU4wptsZyqVLLjWdyYCQpNjR6l4gin91TvULaWx32Ot443V+tNbzU/pcawf04lihCy6KAthfG9EHKlDA/r3ennltv5Ze7C13p+isaGyKo0M5kx4UDvURA4Bxsh70k1caHqwExFszEex8cgZrQ1pPHysk4R7bVjTznAuX/p2vUZ9e2vkZr6xntFkf0YmzvcCE5/KdjZqXh3lMpCde6FPhXrd/jyoYL2iwVf6vLhrBvc2/Crsjq7vQnzrDSvkLsgY=~1; ak_bmsc=A6245E458AADC07EDF98845AAAFA3C7C~000000000000000000000000000000~YAAQpeQ1Fx1NFBSIAQAAX0t2TRNeeWcsOrFN4OdF6K+Vpu7CK4u2FUsb/iGtnMlu1slKbWxbxNrn8vKwKNwcv8ra7IOU8EYMXiBgjnUBJMbgwnmiixxXtdtUT+qZ5jvZhKj/P7PiskS7oyiZ7y3jBb877/8j2ZRY2+bjtTQa6NbGR6PhPnSDoZMvQBIdhnqE/e9Lq/QFIPLM6XqHWhdEetgWG1lbSXovTb9D35KDCeGkH8BoXM93RsbZWcaqWSpHlamsDwCLkBDiuWBEvIEIPhV+WcPeN+E/7QZdjfgXAysq5GFomm1t94v0JxMnkWJU1ExY5XZvkh0yL1hQ3F3ezDu+RbMQhhiZYxWQ/Ybtdzz5eHL+IjIYAWbfdy02wxJ9MLAXfL8eL3gh3PiQuQaVFc0H1JzILO65NF1zXOQ=; _abck=D8A3879F550D48A325858C749ED5EFC6~0~YAAQX3XTF9NHfD6IAQAAhep3TQmZTkA+hlTKR7DWzYzeVThp5wXDUJkti189ffkiW1L6S5k+poawjkqPKkB/fKRKFCHgI0IMa/Tjrbx3dCDSKdsg0wY9jglohkLO7AKd+sjmnMDUiU8P7SPY/g/HEVMTPDWsjvMmP//ZsQ3VJXNn+1iPFDAPjFy1P0sfR8XRza+CItbtnSyDeFhF498vUyitqRGqUiefLxvVzzlv0+5PSCqH/PCELnSoPuT0ZTL6hdPgXPZnAKDlsvGSWG1toYMiXYjjDdwQpCOp7p57dQwrrFlOVHFU8F1kuWUI3gbtwAohAsGhaTJPYUE/2uSbhJUqcZ53qOs0uUhttgDjqMC6JTtxbPGI0MAJ5nGipmXpOPCBTIntKHsxz3kCd0ZsynwvCiNLADTaFNnEp1ajRa4Q+hCg+6Xp~-1~-1~-1; FUN="{\'search\':[{\'reqUrl\':\'/search.pang\',\'isValid\':true}]}"; cto_bundle=nGPxEl9SaG5yR3k1VDRodiUyQlEwSXdiMnI2Nm1rbFFRdEFrdlRWUEd5UEdndVRoQ0lVOCUyRkxaaExiTWFyWDlxcnFWS3RUNiUyRnFqVlRZMFRyNyUyRkRvczczS1hBdEdFZUxEdlJsazklMkJYekZDTG1GOXpha3YzZSUyQjNUWiUyQk9WU2FreGRSd01QVk0ySHJyb1pTTWFmSGV1aiUyRjBNNDlwRFpnJTNEJTNE; baby-isWide=small; bm_sv=ABBAE37C723B86B41ACFEBCDB0962346~YAAQLQ3VF5d/DxWIAQAAZeF6TRP+Ox+yJNm/qqLY7GCh17/bXQvj5sAPtReZTV2iDvIP5QvFy7tJROBe6fuk6uzFRhKKaRp+RvOCJ/dvxf01e6GMYmjX0erdpGLs2JdJWL0RNzuelix7uTdtMhtPvOuORz1e6t7/WMPKIdggQ8Lp3jBxSBeHXms46UXzU47pn77wzBBxqUdC9U5tSOe2oK4rNDPzMCc2Z7bRRbXPWL/PYTgXgQK1sCnae37KkIT45Fo=~1',
        'referer': 'https://www.coupang.com/vp/products/7183570048?itemId=18122029287&vendorItemId=85273125519&pickType=COU_PICK&q=%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8&itemsCount=36&searchId=21451b80c0dd4b08a9569a8a60cdb552&rank=2&isAddedCart=',
        'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
        'x-requested-with': '',
    }

    params = {
        'itemId': str(data['itemId']),
        'selectedId': str(data['vendorItemId']),
    }

    response = requests.get(
        'https://www.coupang.com/vp/products/{}/other-seller-json'.format(data['productNo']),
        params=params,
        cookies=cookies,
        headers=headers,
    )

    results=json.loads(response.text)
    # pprint.pprint(results)
    isOthers=results['totalCount']
    url=""
    if isOthers==0:
        url='https://www.coupang.com/vp/products/{}?itemId={}&vendorItemId={}'.format(data['productNo'],data['itemId'],data['vendorItemId'])
    else:
        results = json.loads(response.text)['items']
        for result in results:
            isRocket=result['deliveryInfo']['badgeType']
            vendorName = result['vendorName']
            print('isRocket:',isRocket,'vendorName:',vendorName)

            if isRocket!="ROCKET" and vendorName!="쿠팡" and vendorName not in prevVendorName:
                vendorItemId=result['vendorItemId']
                vendorName=result['vendorName']
                print(vendorItemId,vendorName)
                url='https://www.coupang.com/vp/products/{}?itemId={}&vendorItemId={}'.format(data['productNo'],data['itemId'],vendorItemId)
                break
    if url=="":
        print("중복되서 업체 없음")


    print("첫업체보이는URL",url)
    regex=re.compile("vendorItemId=\d+")
    changedVendorItemId=regex.findall(url)[0].replace("vendorItemId=","")
    return url,changedVendorItemId,isOthers
def getCompanyCode(data,keyword,changedVendorItemId):
    cookies = {
        'PCID': '7502917043777569504480',
        'x-coupang-accept-language': 'ko-KR',
        'x-coupang-target-market': 'KR',
        '_fbp': 'fb.1.1684062714833.1964233407',
        'sid': '0cd4049032304be58573fada46c29598a51802bc',
        'bm_sz': '40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456',
        'clickCoach': 'yes',
        '_ga': 'GA1.2.1852651095.1684922212',
        '_gid': 'GA1.2.708582323.1684922212',
        'searchKeyword': '%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80',
        'searchKeywordType': '%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D',
        'overrideAbTestGroup': '%5B%5D',
        'MARKETID': '7502917043777569504480',
        'bm_mi': 'ADB92FE5913B40FC5409CA9E81DA620B~YAAQpeQ1FwhMFBSIAQAAq0R2TRP61gInl5atkg8ZN6y2osbd38CFAPZUGFdWAU4wptsZyqVLLjWdyYCQpNjR6l4gin91TvULaWx32Ot443V+tNbzU/pcawf04lihCy6KAthfG9EHKlDA/r3ennltv5Ze7C13p+isaGyKo0M5kx4UDvURA4Bxsh70k1caHqwExFszEex8cgZrQ1pPHysk4R7bVjTznAuX/p2vUZ9e2vkZr6xntFkf0YmzvcCE5/KdjZqXh3lMpCde6FPhXrd/jyoYL2iwVf6vLhrBvc2/Crsjq7vQnzrDSvkLsgY=~1',
        # 'ak_bmsc': 'A6245E458AADC07EDF98845AAAFA3C7C~000000000000000000000000000000~YAAQpeQ1Fx1NFBSIAQAAX0t2TRNeeWcsOrFN4OdF6K+Vpu7CK4u2FUsb/iGtnMlu1slKbWxbxNrn8vKwKNwcv8ra7IOU8EYMXiBgjnUBJMbgwnmiixxXtdtUT+qZ5jvZhKj/P7PiskS7oyiZ7y3jBb877/8j2ZRY2+bjtTQa6NbGR6PhPnSDoZMvQBIdhnqE/e9Lq/QFIPLM6XqHWhdEetgWG1lbSXovTb9D35KDCeGkH8BoXM93RsbZWcaqWSpHlamsDwCLkBDiuWBEvIEIPhV+WcPeN+E/7QZdjfgXAysq5GFomm1t94v0JxMnkWJU1ExY5XZvkh0yL1hQ3F3ezDu+RbMQhhiZYxWQ/Ybtdzz5eHL+IjIYAWbfdy02wxJ9MLAXfL8eL3gh3PiQuQaVFc0H1JzILO65NF1zXOQ=',
        '_abck': 'D8A3879F550D48A325858C749ED5EFC6~0~YAAQX3XTF9NHfD6IAQAAhep3TQmZTkA+hlTKR7DWzYzeVThp5wXDUJkti189ffkiW1L6S5k+poawjkqPKkB/fKRKFCHgI0IMa/Tjrbx3dCDSKdsg0wY9jglohkLO7AKd+sjmnMDUiU8P7SPY/g/HEVMTPDWsjvMmP//ZsQ3VJXNn+1iPFDAPjFy1P0sfR8XRza+CItbtnSyDeFhF498vUyitqRGqUiefLxvVzzlv0+5PSCqH/PCELnSoPuT0ZTL6hdPgXPZnAKDlsvGSWG1toYMiXYjjDdwQpCOp7p57dQwrrFlOVHFU8F1kuWUI3gbtwAohAsGhaTJPYUE/2uSbhJUqcZ53qOs0uUhttgDjqMC6JTtxbPGI0MAJ5nGipmXpOPCBTIntKHsxz3kCd0ZsynwvCiNLADTaFNnEp1ajRa4Q+hCg+6Xp~-1~-1~-1',
        'FUN': '"{\'search\':[{\'reqUrl\':\'/search.pang\',\'isValid\':true}]}"',
        'baby-isWide': 'small',
        'cto_bundle': 'ZgaFg19SaG5yR3k1VDRodiUyQlEwSXdiMnI2NmdLdWVhZFFUSjhsJTJGQWVBM3ZFa29OcXBzeXJxSUdSZ083MEV0WnUydnRvSEk1ZzJmJTJGbjJiaHNUM1QzeHJvVmdJeGp6SHAzMWtwUGpRNk5KTDh2dU41cmRSV2RjYzFDa3c3aGJvZGhXZ2FYOXJkYlhwVEpoN0VGTmtnQ0hicnZnOWclM0QlM0Q',
        'bm_sv': 'ABBAE37C723B86B41ACFEBCDB0962346~YAAQr3XTF0PN7xGIAQAA36B/TRPll9s0SXdFud7mBrdA/XwXCad6fplsSIpeUSXTvJmLjWqjgf2F2Zs7i1c6KRdl78HF48JthQ5jpN9cX4kN3VQHsb/0RgSLA+1W5x9kYn4FeTiVqv9OR1KXQLnl4+smk4Fz/rKfw5CR+IzkhsfbLXGEGenZnTdxYX6Y6AG2LSAorKX2VoczGWQeUkJ65E/ECaVSa1hOH13i1pZcPQzngbzfvovevId0LZF3/mboTfQ=~1',
    }

    headers = {
        'authority': 'www.coupang.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        # 'cookie': 'PCID=7502917043777569504480; x-coupang-accept-language=ko-KR; x-coupang-target-market=KR; _fbp=fb.1.1684062714833.1964233407; sid=0cd4049032304be58573fada46c29598a51802bc; bm_sz=40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456; clickCoach=yes; _ga=GA1.2.1852651095.1684922212; _gid=GA1.2.708582323.1684922212; searchKeyword=%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80; searchKeywordType=%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D; overrideAbTestGroup=%5B%5D; MARKETID=7502917043777569504480; bm_mi=ADB92FE5913B40FC5409CA9E81DA620B~YAAQpeQ1FwhMFBSIAQAAq0R2TRP61gInl5atkg8ZN6y2osbd38CFAPZUGFdWAU4wptsZyqVLLjWdyYCQpNjR6l4gin91TvULaWx32Ot443V+tNbzU/pcawf04lihCy6KAthfG9EHKlDA/r3ennltv5Ze7C13p+isaGyKo0M5kx4UDvURA4Bxsh70k1caHqwExFszEex8cgZrQ1pPHysk4R7bVjTznAuX/p2vUZ9e2vkZr6xntFkf0YmzvcCE5/KdjZqXh3lMpCde6FPhXrd/jyoYL2iwVf6vLhrBvc2/Crsjq7vQnzrDSvkLsgY=~1; ak_bmsc=A6245E458AADC07EDF98845AAAFA3C7C~000000000000000000000000000000~YAAQpeQ1Fx1NFBSIAQAAX0t2TRNeeWcsOrFN4OdF6K+Vpu7CK4u2FUsb/iGtnMlu1slKbWxbxNrn8vKwKNwcv8ra7IOU8EYMXiBgjnUBJMbgwnmiixxXtdtUT+qZ5jvZhKj/P7PiskS7oyiZ7y3jBb877/8j2ZRY2+bjtTQa6NbGR6PhPnSDoZMvQBIdhnqE/e9Lq/QFIPLM6XqHWhdEetgWG1lbSXovTb9D35KDCeGkH8BoXM93RsbZWcaqWSpHlamsDwCLkBDiuWBEvIEIPhV+WcPeN+E/7QZdjfgXAysq5GFomm1t94v0JxMnkWJU1ExY5XZvkh0yL1hQ3F3ezDu+RbMQhhiZYxWQ/Ybtdzz5eHL+IjIYAWbfdy02wxJ9MLAXfL8eL3gh3PiQuQaVFc0H1JzILO65NF1zXOQ=; _abck=D8A3879F550D48A325858C749ED5EFC6~0~YAAQX3XTF9NHfD6IAQAAhep3TQmZTkA+hlTKR7DWzYzeVThp5wXDUJkti189ffkiW1L6S5k+poawjkqPKkB/fKRKFCHgI0IMa/Tjrbx3dCDSKdsg0wY9jglohkLO7AKd+sjmnMDUiU8P7SPY/g/HEVMTPDWsjvMmP//ZsQ3VJXNn+1iPFDAPjFy1P0sfR8XRza+CItbtnSyDeFhF498vUyitqRGqUiefLxvVzzlv0+5PSCqH/PCELnSoPuT0ZTL6hdPgXPZnAKDlsvGSWG1toYMiXYjjDdwQpCOp7p57dQwrrFlOVHFU8F1kuWUI3gbtwAohAsGhaTJPYUE/2uSbhJUqcZ53qOs0uUhttgDjqMC6JTtxbPGI0MAJ5nGipmXpOPCBTIntKHsxz3kCd0ZsynwvCiNLADTaFNnEp1ajRa4Q+hCg+6Xp~-1~-1~-1; FUN="{\'search\':[{\'reqUrl\':\'/search.pang\',\'isValid\':true}]}"; baby-isWide=small; cto_bundle=ZgaFg19SaG5yR3k1VDRodiUyQlEwSXdiMnI2NmdLdWVhZFFUSjhsJTJGQWVBM3ZFa29OcXBzeXJxSUdSZ083MEV0WnUydnRvSEk1ZzJmJTJGbjJiaHNUM1QzeHJvVmdJeGp6SHAzMWtwUGpRNk5KTDh2dU41cmRSV2RjYzFDa3c3aGJvZGhXZ2FYOXJkYlhwVEpoN0VGTmtnQ0hicnZnOWclM0QlM0Q; bm_sv=ABBAE37C723B86B41ACFEBCDB0962346~YAAQr3XTF0PN7xGIAQAA36B/TRPll9s0SXdFud7mBrdA/XwXCad6fplsSIpeUSXTvJmLjWqjgf2F2Zs7i1c6KRdl78HF48JthQ5jpN9cX4kN3VQHsb/0RgSLA+1W5x9kYn4FeTiVqv9OR1KXQLnl4+smk4Fz/rKfw5CR+IzkhsfbLXGEGenZnTdxYX6Y6AG2LSAorKX2VoczGWQeUkJ65E/ECaVSa1hOH13i1pZcPQzngbzfvovevId0LZF3/mboTfQ=~1',
        'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
    }

    # params = {
    #     'itemId': str(data['itemId']),
    #     'vendorItemId': str(changedVendorItemId),
    #     'q': str(keyword),
    # }
    # print(data['itemId'])
    # print(changedVendorItemId)
    url="https://www.coupang.com/vp/products/{}?itemId={}&vendorItemId={}&isAddedCart=".format(data['productNo'],data['itemId'],changedVendorItemId)
    response = requests.get(url, cookies=cookies,
                            headers=headers)

    soup=BeautifulSoup(response.text,'lxml')
    # print(soup.prettify())
    vendorInfo=soup.find('a',attrs={'class':'prod-sale-vendor-name'})['href']
    vendorName = soup.find('a', attrs={'class': 'prod-sale-vendor-name'}).get_text()
    print('vendorInfo:',vendorInfo)
    print('vendorName:', vendorName)
    regex=re.compile("A\d+")
    vendorId=regex.findall(vendorInfo)[0]
    print("vendorId",vendorId)
    return vendorId,vendorName
def getCompanyProductList(vendorId,companyInfo,keyword,vendorName,isOthers):
    productList=[]
    cookies = {
        'PCID': '7502917043777569504480',
        'x-coupang-accept-language': 'ko-KR',
        'x-coupang-target-market': 'KR',
        '_fbp': 'fb.1.1684062714833.1964233407',
        'sid': '0cd4049032304be58573fada46c29598a51802bc',
        'MARKETID': '7502917043777569504480',
        'FUN': '"{\'search\':[{\'reqUrl\':\'/search.pang\',\'isValid\':true}]}"',
        'bm_sz': '40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456',
        # 'ak_bmsc': 'B41F73EC24F90E77CF738C624CF14779~000000000000000000000000000000~YAAQtAI1Fw0meDGIAQAAqbICTROVsA0V/qudWRiSwnljg43O1oGWjJPUEijCnC5pUsqeuzoR1FaD1p9coDKixPSZcz1MaMgqxbUq71GMGoU6QJq1gVIQx60GHEKfy41kF/O1O+0bOSUErLRmIP9HdZvQpEL44nh9dMFtcjVf6Vq/zYFjs9UuzttJteBL7QjB1MzmC9p85pg2JVqyOS+hcmfbnlr+B499GRQtjpcPgPKgQ7sbVfIDhxOrYhGN8LDwzQ0nbus0AtHf5mKxdBA3QFh2Ln7X6YI22pflqyJqLWnqBA+nh06FBu02Gu0vKrmuyOvy/GTXC1GvUiNPHqm8hlfxTq0+qKbTIW3okKFgC5BOPL/TvqCYeO5LdYtNPZmnROjMCqNInDsL8yhXUsj/sRgcpnFQ/3ypsokCSyjkYJPTdzQEuJsTZsdUsd7ovDAhtczX/8mwTwFa3Wo5uOUkfLnKpE1DvDcil0PcyB5AHqIUhXsSmciKL96H',
        'clickCoach': 'yes',
        '_ga': 'GA1.2.1852651095.1684922212',
        '_gid': 'GA1.2.708582323.1684922212',
        '_abck': 'D8A3879F550D48A325858C749ED5EFC6~0~YAAQRiPJFx0bdDeIAQAAD84xTQkMzsWq+PhFIdU84itWHyNTwcXl+ytdZav7BHIiP4xSwP4nlC6jFHUmwuGZDJX6nD7U6+2IiFjkQHBnSpXwBelghEI2IytXVY3BoC3Em0Z6NOm8F06HA4EmNCPtFNiWbNFisCd7w+xuuiRZL9A9t49nNBZqdjhWD7fAW29KQwCaOAeqvYcPpsaqQ+ym9s5gNVXh2yLYZBQlADdsLtP3Jl22T2mHGNnoxlEjvTPIfLeuPDXTzdMIcsu+VHyz9oHuru9UWY8lhDMyGkK7MmYSOqlXwPib7u6NKZvVyundu3ZlaM73BdoZamua62MdR7QDaL32uLriASGWlYhwJk2wh88bh8YhK6Te95ZH4J39wzZdu8ipMY6PEIdlaKkdbBR83l9l7UH/fMQRMzym7OWnwIo8rj+3~-1~-1~-1',
        'searchKeyword': '%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80',
        'searchKeywordType': '%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D',
        'overrideAbTestGroup': '%5B%5D',
        'cto_bundle': 'P8k1BF9SaG5yR3k1VDRodiUyQlEwSXdiMnI2NmwlMkJVS3hidHRTRzZwVnNJUHNqb094ZFBLMk5sYzlqQVladUZHbkY2WGk2dnJCZUV6R2o5OWowRzR0Rlg4bSUyQk9lJTJGVGhtdjJVSzg2UXZNMmQzMzJQdElPQUhDaWZZeFowZEJsclpzSmJnYWlQQUJiTXFVVXBuZmxDVzVUUjJ4MnNwUSUzRCUzRA',
        'bm_sv': 'F8BF408D3FAE4A31AB9E95EF1549BE87~YAAQXyPJFyMCqxCIAQAARJtETROYL4AozA7/C5lZWKoeq03px/R/hs/nUwJiJ2KcOF53e4TsaE+LI/sv46ZD+qBQvr+5DrMIpLT6Pab6cCZYGmx6agYsZAR9TA0AFvYqY+yt4iAF1ytLfpg/0yBX9JVgoHBYTRoPhwfKoYkYkPq4Bg98stFFWoP0bwAMRPyOV51nofrkXZvkc78TtOhnOYRostBtjsxcJVgbHvBS+cqYF8gaX2+dvr3JGysgRhSSi/60~1',
        'baby-isWide': 'small',
    }

    headers = {
        'authority': 'store.coupang.com',
        'accept': 'application/json, text/javascript, */*; q=0.01',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'content-type': 'application/json; charset=utf-8',
        # 'cookie': 'PCID=7502917043777569504480; x-coupang-accept-language=ko-KR; x-coupang-target-market=KR; _fbp=fb.1.1684062714833.1964233407; sid=0cd4049032304be58573fada46c29598a51802bc; MARKETID=7502917043777569504480; FUN="{\'search\':[{\'reqUrl\':\'/search.pang\',\'isValid\':true}]}"; bm_sz=40EF2C4CBEB2C6D9599A1EFBEBAE4FC5~YAAQtAI1F38leDGIAQAA7a4CTROxxUjTbkEDHLJuKs90Zd7UtCNdGqaXgGSdMQXJFk/cQVEZy4Q5pKtkV5pJN8WjTuqFP9p1ypjLqcPM9ANctLXrVrYd/f/U6/0cRb8u1UGJaBG+xmGBrshwEWDy0o6lT9fMIHgj3FYE1XLUtTDj+r01sa9vzRJHVxgOHgTDuo8Yiz2Yy2N1a4ItS5u1mR1qU7FPDIA/8mYGHamsqEgSCVTjY2R2F2tkOAokYg14Ez42sXrcx2flLHawglMhsOOhmld4w7+yhd3dPz+VcewYv6fk~3421240~4273456; ak_bmsc=B41F73EC24F90E77CF738C624CF14779~000000000000000000000000000000~YAAQtAI1Fw0meDGIAQAAqbICTROVsA0V/qudWRiSwnljg43O1oGWjJPUEijCnC5pUsqeuzoR1FaD1p9coDKixPSZcz1MaMgqxbUq71GMGoU6QJq1gVIQx60GHEKfy41kF/O1O+0bOSUErLRmIP9HdZvQpEL44nh9dMFtcjVf6Vq/zYFjs9UuzttJteBL7QjB1MzmC9p85pg2JVqyOS+hcmfbnlr+B499GRQtjpcPgPKgQ7sbVfIDhxOrYhGN8LDwzQ0nbus0AtHf5mKxdBA3QFh2Ln7X6YI22pflqyJqLWnqBA+nh06FBu02Gu0vKrmuyOvy/GTXC1GvUiNPHqm8hlfxTq0+qKbTIW3okKFgC5BOPL/TvqCYeO5LdYtNPZmnROjMCqNInDsL8yhXUsj/sRgcpnFQ/3ypsokCSyjkYJPTdzQEuJsTZsdUsd7ovDAhtczX/8mwTwFa3Wo5uOUkfLnKpE1DvDcil0PcyB5AHqIUhXsSmciKL96H; clickCoach=yes; _ga=GA1.2.1852651095.1684922212; _gid=GA1.2.708582323.1684922212; _abck=D8A3879F550D48A325858C749ED5EFC6~0~YAAQRiPJFx0bdDeIAQAAD84xTQkMzsWq+PhFIdU84itWHyNTwcXl+ytdZav7BHIiP4xSwP4nlC6jFHUmwuGZDJX6nD7U6+2IiFjkQHBnSpXwBelghEI2IytXVY3BoC3Em0Z6NOm8F06HA4EmNCPtFNiWbNFisCd7w+xuuiRZL9A9t49nNBZqdjhWD7fAW29KQwCaOAeqvYcPpsaqQ+ym9s5gNVXh2yLYZBQlADdsLtP3Jl22T2mHGNnoxlEjvTPIfLeuPDXTzdMIcsu+VHyz9oHuru9UWY8lhDMyGkK7MmYSOqlXwPib7u6NKZvVyundu3ZlaM73BdoZamua62MdR7QDaL32uLriASGWlYhwJk2wh88bh8YhK6Te95ZH4J39wzZdu8ipMY6PEIdlaKkdbBR83l9l7UH/fMQRMzym7OWnwIo8rj+3~-1~-1~-1; searchKeyword=%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%7C%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%7C%EA%B3%BC%EC%9E%90%7C%EC%9C%84%EC%85%80; searchKeywordType=%7B%22%EB%B0%94%EC%9D%B4%EC%98%A4%ED%8B%B4%ED%83%88%EB%AA%A8%EC%83%B4%ED%91%B8%22%3A0%7D%7C%7B%22%EC%BD%94%EC%B9%B4%EC%BD%9C%EB%9D%BC%22%3A0%7D%7C%7B%22%EA%B3%BC%EC%9E%90%22%3A0%7D%7C%7B%22%EC%9C%84%EC%85%80%22%3A0%7D; overrideAbTestGroup=%5B%5D; cto_bundle=P8k1BF9SaG5yR3k1VDRodiUyQlEwSXdiMnI2NmwlMkJVS3hidHRTRzZwVnNJUHNqb094ZFBLMk5sYzlqQVladUZHbkY2WGk2dnJCZUV6R2o5OWowRzR0Rlg4bSUyQk9lJTJGVGhtdjJVSzg2UXZNMmQzMzJQdElPQUhDaWZZeFowZEJsclpzSmJnYWlQQUJiTXFVVXBuZmxDVzVUUjJ4MnNwUSUzRCUzRA; bm_sv=F8BF408D3FAE4A31AB9E95EF1549BE87~YAAQXyPJFyMCqxCIAQAARJtETROYL4AozA7/C5lZWKoeq03px/R/hs/nUwJiJ2KcOF53e4TsaE+LI/sv46ZD+qBQvr+5DrMIpLT6Pab6cCZYGmx6agYsZAR9TA0AFvYqY+yt4iAF1ytLfpg/0yBX9JVgoHBYTRoPhwfKoYkYkPq4Bg98stFFWoP0bwAMRPyOV51nofrkXZvkc78TtOhnOYRostBtjsxcJVgbHvBS+cqYF8gaX2+dvr3JGysgRhSSi/60~1; baby-isWide=small',
        'referer': 'https://store.coupang.com/vp/vendors/A00669316/products?vendorName=%EC%97%A0%ED%8C%8C%ED%8A%B8%EB%84%88%EC%8A%A4&productId=7353298574&outboundShippingPlaceId=',
        'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
    }

    params = {
        'outboundShippingPlaceId': '',
        'attributeFilters': '',
        'brand': '',
        'componentId': '',
        'keyword': '',
        'maxPrice': '',
        'minPrice': '',
        'pageNum': '1',
        'rating': '0',
        'sortTypeValue': 'BEST_SELLING',
        'scpLanding': 'false',
    }

    response = requests.get(
        'https://store.coupang.com/vp/vendors/{}/product/lists'.format(vendorId),
        params=params,
        cookies=cookies,
        headers=headers,
    )
    results=json.loads(response.text)['data']['products']
    # pprint.pprint(results)
    for index,result in enumerate(results):
        # pprint.pprint(result)
        if index>=20:
            break

        title=result['title']
        salePrice=result['salePrice']
        reviewRatingCount=result['reviewRatingCount']
        reviewRatingAverage=result['reviewRatingAverage']
        realPrice=round(salePrice*0.8,-3)
        promiseDeliveryDate=result['promiseDeliveryDate']
        url=result['link']
        # print('promiseDeliveryDate:',promiseDeliveryDate)
        try:
            regex=re.compile("\d+/\d+")
            promiseDeliveryDateChanged=regex.findall(promiseDeliveryDate)[0]
        except:
            print("배송기간 부재로 기존 배송일 사용")
        timeNow=datetime.datetime.now().strftime("%m/%d")
        try:
            duration=calculate_duration(timeNow,promiseDeliveryDateChanged)
        except:
            duration="X"

        data={'keyword':keyword,'vendorName':vendorName,'rating':index+1,'title':title,'salePrice':salePrice,'realPrice':realPrice,'reviewRatingCount':reviewRatingCount,'reviewRatingAverage':reviewRatingAverage,'duration':duration,'isOthers':isOthers+1,'url':url}
        productList.append(data)
    # print("productList:",productList)
    return productList


wb=openpyxl.Workbook()
ws=wb.active
# ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10) #저장할 셀
columnName=['검색키워드','스토어명','판매량순위','상품명','상품가','상품가*0.8','리뷰수','판매자수','별점','배송기간']
ws.column_dimensions['D'].width=40 # D열 사이즈 저종
ws.append(['','','','','','','','','',''])
ws.append(['','','','카페 바로가기','','','','','',''])
# ws.cell(row=2, column=3).value = '↑↑↑'
ws.cell(row=2, column=4).hyperlink = 'https://cafe.naver.com/dbs1m1h' # 카페 바로가기
ws.cell(row=2, column=4).style = "Hyperlink"
ws.cell(row=2,column=4).alignment=Alignment(horizontal='center')
for i in range(1,11):
    ws.cell(row=3,column=i).alignment=Alignment(horizontal='center')
    ws.cell(row=3, column=i).fill = PatternFill(fill_type="solid",fgColor=Color('789ABC'))
    ws.cell(row=3,column=i).value=columnName[i-1]



## 이미지 불러오기
image_path = 'banner.png'
image = Image(image_path)

ws.add_image(image, 'D1')  ## 이미지 삽입

## 이미지 픽셀을 셀 폭과 높이로 변환
col_width, row_height = get_col_width_row_height(image.width, image.height)

# ws.column_dimensions['A'].width = col_width  ## 셀 폭 변경
ws.row_dimensions[1].height = row_height  ## 셀 높이 변경

keywordList=load_excel('keyword.xlsx')
timeNow=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
for keyword in keywordList:
    dataList=getUrlList(keyword)
    if len(dataList)==0:
        print("상품없어서스킵함")
        continue
    productList=[]
    count=0
    prevVendorName=[]
    for index,dataElem in enumerate(dataList):
        if count>=5:
            break
        companyInfo=dataElem
        time.sleep(random.randint(5,10)*0.1)
        print('companyInfo:',companyInfo,'keyword:',keyword)
        try:
            url,changedVendorItemId,isOthers=getCompanyInfo(companyInfo,prevVendorName)
        except:
            print("업체중복으로안됨")
            continue
        time.sleep(random.randint(5, 10) * 0.1)
        try:
            vendorId,vendorName=getCompanyCode(dataElem,keyword,changedVendorItemId)
            prevVendorName.append(vendorName)
        except:
            print("로켓배송임")
            continue
        time.sleep(random.randint(5, 10) * 0.1)
        dataList=getCompanyProductList(vendorId,companyInfo,keyword,vendorName,isOthers)
        time.sleep(random.randint(5, 10) * 0.1)
        productList.extend(dataList)
        time.sleep(random.randint(10,15)*0.1)
        count=count+1
    count=4
    for productElem in productList:
        ws.cell(row=count,column=1).value=productElem['keyword']
        ws.cell(row=count,column=2).value=productElem['vendorName']
        ws.cell(row=count, column=3).value = productElem['rating']
        ws.cell(row=count, column=4).value = productElem['title']
        ws.cell(row=count, column=4).hyperlink = productElem['url']
        ws.cell(row=count, column=4).style = "Hyperlink"
        ws.cell(row=count, column=5).value = productElem['salePrice']
        ws.cell(row=count, column=6).value = productElem['realPrice']
        ws.cell(row=count, column=7).value = productElem['reviewRatingCount']
        ws.cell(row=count, column=8).value = productElem['isOthers']
        ws.cell(row=count, column=9).value = productElem['reviewRatingAverage']
        ws.cell(row=count, column=10).value = productElem['duration']




        # data=[productElem['keyword'],productElem['vendorName'],productElem['rating'],productElem['title'],productElem['salePrice'],productElem['realPrice'],
        #       productElem['reviewRatingCount'],productElem['isOthers'],productElem['reviewRatingAverage'],productElem['duration']]
        # ws.append(data)
        count=count+1

    first = FormatObject(type='min')
    second = FormatObject(type='max')

    data_bar = DataBar(cfvo=[first, second], color="ADD8E6",
                       showValue=None, minLength=None, maxLength=None)

    rule = Rule(type='dataBar', dataBar=data_bar)
    target_column = 3  ## 네 번째 칼럼
    column_letter = get_column_letter(target_column)  ## D
    start_row = ws[column_letter][1].coordinate  ## D2
    end_row = ws[column_letter][-1].coordinate  ## D17

    for x in ws[start_row:end_row]:
        try:
            cell = x[0]
            cell.value = int(cell.value)
        except:
            print("없음")
    ws.conditional_formatting.add(f"{start_row}:{end_row}", rule)

    first = FormatObject(type='min')
    second = FormatObject(type='max')

    data_bar = DataBar(cfvo=[first, second], color="ADD8E6",
                       showValue=None, minLength=None, maxLength=None)

    rule = Rule(type='dataBar', dataBar=data_bar)
    target_column = 5  ## 네 번째 칼럼
    column_letter = get_column_letter(target_column)  ## D
    start_row = ws[column_letter][1].coordinate  ## D2
    end_row = ws[column_letter][-1].coordinate  ## D17

    for x in ws[start_row:end_row]:
        try:
            cell = x[0]
            cell.value = int(cell.value)
        except:
            print("없음")
    ws.conditional_formatting.add(f"{start_row}:{end_row}", rule)

    wb.save('result_{}.xlsx'.format(timeNow))


